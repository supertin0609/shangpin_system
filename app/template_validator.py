from pathlib import Path
import re

from openpyxl import load_workbook

from .paths import OUTPUTS_DIR
from .listing_rules import validate_listing_row
from .template_sheet import find_template_sheet, template_sheet_names_text
from .variation_title_rules import (
    PARENT_COLOR_PHRASE,
    PARENT_STYLE_PHRASE,
    has_parent_only_phrase,
    package_prefix_for_row,
    title_starts_with_package_prefix,
)


FIELD_NAMES = {
    "sku": "contribution_sku#1.value",
    "skip_offer": "skip_offer[marketplace_id=ATVPDKIKX0DER]#1.value",
    "condition": "condition_type[marketplace_id=ATVPDKIKX0DER]#1.value",
    "list_price": "list_price[marketplace_id=ATVPDKIKX0DER]#1.value",
    "haul_price": "purchasable_offer[marketplace_id=ATVPDKIKX0DER][audience=BZR]#1.our_price#1.schedule#1.value_with_tax",
    "minimum_seller_allowed_price": "purchasable_offer[marketplace_id=ATVPDKIKX0DER][audience=BZR]#1.minimum_seller_allowed_price#1.schedule#1.value_with_tax",
    "item_depth": "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.depth.value",
    "item_depth_unit": "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.depth.unit",
    "item_height": "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.value",
    "item_height_unit": "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
    "item_width": "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.value",
    "item_width_unit": "item_depth_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
    "parentage_level": "parentage_level[marketplace_id=ATVPDKIKX0DER]#1.value",
    "parent_sku": "child_parent_sku_relationship[marketplace_id=ATVPDKIKX0DER]#1.parent_sku",
    "variation_theme": "variation_theme#1.name",
    "title": "item_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "item_highlight": "title_differentiation[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "description": "product_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "generic_keyword": "generic_keyword[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "color": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "size": "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "number_of_items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
    "item_package_quantity": "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value",
    "contains_battery_or_cell": "contains_battery_or_cell[marketplace_id=ATVPDKIKX0DER]#1.value",
    "dog_breed_size": "dog_breed_size[marketplace_id=ATVPDKIKX0DER]#1.value",
}

CJK_RE = re.compile(r"[\u4e00-\u9fff]")


def _text(value):
    return str(value or "").strip()


def _contains_exact_text(text, value):
    text = _text(text)
    value = _text(value)
    if not text or not value:
        return False
    return re.search(rf"(?<![A-Za-z0-9]){re.escape(value)}(?![A-Za-z0-9])", text, re.I) is not None


def _unique_values(items, key):
    values = []
    seen = set()
    for item in items:
        value = _text(item.get(key))
        if not value:
            continue
        lowered = value.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        values.append(value)
    return values


def _valid_values_row(wb, label_prefix):
    if "Valid Values" not in wb.sheetnames:
        return set()
    ws = wb["Valid Values"]
    for row in ws.iter_rows(values_only=True):
        values = [_text(value) for value in row]
        if any(value.startswith(label_prefix) for value in values):
            return {value for value in values if value and not value.startswith(label_prefix)}
    return set()


def _variation_theme_values(wb, product_type):
    product_type = _text(product_type)
    if not product_type:
        return set()
    return _valid_values_row(wb, f"Variation Theme Name - [ {product_type} ]")


def _requires_package_prefix(info):
    prefix = package_prefix_for_row({
        "set_count": info.get("package_quantity") or info.get("number_of_items"),
        "title": "",
    })
    return bool(prefix)


def _validate_variation_titles(findings, row_infos):
    parents_by_sku = {
        _text(info.get("sku")): info
        for info in row_infos.values()
        if _text(info.get("parentage")).lower() == "parent" and _text(info.get("sku"))
    }
    children_by_parent = {}
    for info in row_infos.values():
        if _text(info.get("parentage")).lower() != "child":
            continue
        parent_sku = _text(info.get("parent_sku"))
        if parent_sku:
            children_by_parent.setdefault(parent_sku, []).append(info)

    for info in row_infos.values():
        title = _text(info.get("title"))
        sku = info.get("sku")
        if _requires_package_prefix(info) and not title_starts_with_package_prefix(title):
            findings.append(error(info["row"], "Title", f"{sku} 是多件装/套装，但标题未以数量或套装前缀开头。", "多件装和套装标题首位写明数量，例如 50 Pcs 或 2 Set，然后再接商品关键词。"))
        if _text(info.get("parentage")).lower() == "child":
            if has_parent_only_phrase(title):
                findings.append(error(info["row"], "Title", f"{sku} 是 Child 行，但标题包含父体总结词。", f"Child 标题只保留通用标题和自身变体属性，不填写 {PARENT_COLOR_PHRASE} 或 {PARENT_STYLE_PHRASE}。"))

    for parent_sku, children in children_by_parent.items():
        parent = parents_by_sku.get(parent_sku)
        if not parent:
            continue
        parent_title = _text(parent.get("title"))
        theme = " ".join(_text(item.get("variation_theme")).lower() for item in [parent] + children)
        colors = _unique_values(children, "color")
        sizes = _unique_values(children, "size")
        if len(colors) > 1 and not _contains_exact_text(parent_title, PARENT_COLOR_PHRASE):
            findings.append(error(parent["row"], "Title", f"{parent_sku} 是多颜色 Parent，但标题未说明多色可选。", f"Parent 标题基于通用标题生成，并加入 {PARENT_COLOR_PHRASE}。"))
        if len(sizes) > 1 and not _contains_exact_text(parent_title, PARENT_STYLE_PHRASE):
            findings.append(error(parent["row"], "Title", f"{parent_sku} 是多款式 Parent，但标题未说明多款式可选。", f"Parent 标题基于通用标题生成，并加入 {PARENT_STYLE_PHRASE}。"))

        for child in children:
            child_title = _text(child.get("title"))
            if "color" in theme and _text(child.get("color")) and not _contains_exact_text(child_title, child.get("color")):
                findings.append(error(child["row"], "Title", f"{child.get('sku')} 是颜色变体 Child，但标题未包含自身颜色 `{child.get('color')}`。", "Child 标题在通用标题基础上追加自己的颜色属性。"))
            if "size" in theme and _text(child.get("size")) and not _contains_exact_text(child_title, child.get("size")):
                findings.append(error(child["row"], "Title", f"{child.get('sku')} 是款式/尺寸变体 Child，但标题未包含自身属性 `{child.get('size')}`。", "Child 标题在通用标题基础上追加自己的尺寸或款式属性。"))


COPY_FIELD_NAMES = {
    "Title": FIELD_NAMES["title"],
    "Description": FIELD_NAMES["description"],
    "Bullet 1": "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    "Bullet 2": "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#2.value",
    "Bullet 3": "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#3.value",
    "Bullet 4": "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#4.value",
    "Bullet 5": "bullet_point[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#5.value",
}


PRODUCT_TYPE_CONDITIONAL_FIELDS = {
    "PLANTER": {
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Special Features": "special_feature[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Mounting Type": "mounting_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Product Compliance Certificate": "required_product_compliance_certificate[marketplace_id=ATVPDKIKX0DER]#1.value",
    },
    "HEALTH_PERSONAL_CARE": {
        "Is the Item Heat Sensitive?": "is_heat_sensitive[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Is Product Expirable": "is_expiration_dated_product[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Unit Count": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Unit Count Type": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US].value",
        "Contains Liquid Contents?": "contains_liquid_contents[marketplace_id=ATVPDKIKX0DER]#1.value",
    },
    "ANIMAL_COLLAR": {
        "Item Type Keyword": "item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Number": "model_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Manufacturer": "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Style": "style[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Material": "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Number of Items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Package Quantity": "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Color Map": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.standardized_values#1",
        "Color": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Size": "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Part Number": "part_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Care Instructions": "care_instructions[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Pattern": "pattern[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Unit Count": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Unit Count Type": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US].value",
        "Included Components": "included_components[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Specific Uses for Product": "specific_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Closure Type": "closure[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US]#1.value",
        "Item Length Longer Side": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Item Length Unit": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Width Shorter Side": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Item Width Unit": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Dog Breed Size": "dog_breed_size[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Number of Packs": "number_of_packs[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Pet Type": "pet_type[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Animal Collar Type": "animal_collar_type[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight Unit": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Product Tax Code": "product_tax_code#1.value",
        "Item Package Length": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Package Length Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Package Width": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Package Width Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Package Height": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Package Height Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Package Weight": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Package Weight Unit": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Number of Boxes": "number_of_boxes[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Country of Origin": "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries required?": "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries included?": "batteries_included[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Dangerous Goods Regulations": "supplier_declared_dg_hz_regulation[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Directions": "directions[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    },
    "BOTTLE": {
        "Item Type Keyword": "item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Number": "model_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Manufacturer": "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Special Features": "special_feature[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Style": "style[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Material": "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Number of Items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Package Quantity": "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Color": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Size": "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Part Number": "part_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Shape": "item_shape[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Theme": "theme[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Care Instructions": "care_instructions[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Capacity": "capacity[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Capacity Unit": "capacity[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Volume Capacity": "volume_capacity_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Volume Capacity Unit": "volume_capacity_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.unit",
        "Pattern": "pattern[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Unit Count": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Unit Count Type": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US].value",
        "Included Components": "included_components[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Specific Uses For Product": "specific_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Recommended Uses For Product": "recommended_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Bottle Color": "bottle[marketplace_id=ATVPDKIKX0DER]#1.color[language_tag=en_US]#1.value",
        "Reusability": "reusability[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Item Height Base to Top": "item_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Item Height Unit (Base to Top)": "item_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Item Width Top": "item_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Item Width Unit (Top)": "item_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Number of Packs": "number_of_packs[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight Unit": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Product Tax Code": "product_tax_code#1.value",
        "Item Length": "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Item Length Unit": "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Width": "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Item Width Unit": "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Height": "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Item Height Unit": "item_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Item Package Length": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Package Length Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Package Width": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Package Width Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Package Height": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Package Height Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Package Weight": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Package Weight Unit": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Item Display Weight": "item_display_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Display Weight Unit": "item_display_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Country of Origin": "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries required?": "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries included?": "batteries_included[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Dangerous Goods Regulations": "supplier_declared_dg_hz_regulation[marketplace_id=ATVPDKIKX0DER]#1.value",
    },
    "COSMETIC_CASE": {
        "Item Type Keyword": "item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Target Audience": "target_audience[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Model Number": "model_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Manufacturer": "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Style": "style[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Department Name": "department[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Target Gender": "target_gender[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Material Type": "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Number of Items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Package Quantity": "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Water Resistance Level": "water_resistance_level[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Color Map": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.standardized_values#1",
        "Color": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Size": "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Part Number": "part_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Shape": "item_shape[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Form Factor": "form_factor[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Pattern": "pattern[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Unit Count": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Unit Count Type": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US].value",
        "Recommended Uses For Product": "recommended_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Closure Type": "closure[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US]#1.value",
        "Height base to top": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Height Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Length longer horizontal edge": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Length Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Width shorter horizontal edge": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Width Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Number of Packs": "number_of_packs[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight Unit": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Product Tax Code": "product_tax_code#1.value",
        "Item Package Length": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Package Length Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Package Width": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Package Width Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Package Height": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Package Height Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Package Weight": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Package Weight Unit": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Country of Origin": "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries required?": "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries included?": "batteries_included[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Dangerous Goods Regulations": "supplier_declared_dg_hz_regulation[marketplace_id=ATVPDKIKX0DER]#1.value",
    },
    "CLEANING_BRUSH": {
        "Item Type Keyword": "item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Package Level": "package_level[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Number": "model_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Manufacturer": "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Style": "style[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Material": "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Number of Items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Package Quantity": "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Color": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Size": "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Part Number": "part_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Shape": "item_shape[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Surface Recommendation": "surface_recommendation[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Handle Material": "handle[marketplace_id=ATVPDKIKX0DER]#1.material[language_tag=en_US]#1.value",
        "Item Firmness Description": "item_firmness_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Pattern": "pattern[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Unit Count": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Unit Count Type": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US].value",
        "Specific Uses for Product": "specific_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Bristle Material": "bristle[marketplace_id=ATVPDKIKX0DER]#1.material[language_tag=en_US]#1.value",
        "Height base to top": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Height Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Length longer horizontal edge": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Length Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Width shorter horizontal edge": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Width Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Width": "item_width[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Width Unit": "item_width[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Item Length": "item_length[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Length Unit": "item_length[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Number of Packs": "number_of_packs[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight Unit": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Product Tax Code": "product_tax_code#1.value",
        "Item Package Length": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Package Length Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Package Width": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Package Width Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Package Height": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Package Height Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Package Weight": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Package Weight Unit": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Country of Origin": "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries required?": "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries included?": "batteries_included[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Dangerous Goods Regulations": "supplier_declared_dg_hz_regulation[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Product Compliance Certificate": "required_product_compliance_certificate[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Compliance - Brush Intended Use": "compliance_brush_intended_use[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Compliance - Is Motorized": "compliance_is_motorized[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Compliance - Is Mechanical": "compliance_is_mechanical[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Compliance - Bristle Material": "compliance_bristle_material[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Compliance - Is Hand-Operated": "compliance_is_hand_operated[marketplace_id=ATVPDKIKX0DER]#1.value",
    },
    "PET_TOY": {
        "Item Type Keyword": "item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Number": "model_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Manufacturer": "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Material": "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Number of Items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Package Quantity": "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Subject Character": "subject_character[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Color": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Size": "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Part Number": "part_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Theme": "theme[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Breed Recommendation": "breed_recommendation[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Unit Count": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Unit Count Type": "unit_count[marketplace_id=ATVPDKIKX0DER]#1.type[language_tag=en_US].value",
        "Included Components": "included_components[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Specific Uses for Product": "specific_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Recommended Uses For Product": "recommended_uses_for_product[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Indoor Outdoor Usage": "indoor_outdoor_usage[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Height base to top": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Height Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Length longer horizontal edge": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Length Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Width shorter horizontal edge": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Width Unit": "item_length_width_height[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Pet Toy Type": "pet_toy_type[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Number of Packs": "number_of_packs[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Pet Type": "pet_type[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight Unit": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Product Tax Code": "product_tax_code#1.value",
        "Item Package Length": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Package Length Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Package Width": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Package Width Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Package Height": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Package Height Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Package Weight": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Package Weight Unit": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Number of Boxes": "number_of_boxes[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Country of Origin": "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries required?": "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries included?": "batteries_included[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Dangerous Goods Regulations": "supplier_declared_dg_hz_regulation[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Contains Liquid Contents?": "contains_liquid_contents[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Directions": "directions[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    },
    "TOWEL": {
        "Model Number": "model_number[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Manufacturer": "manufacturer[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Product Tax Code": "product_tax_code#1.value",
        "Material": "material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Fabric Type": "fabric_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Number of Items": "number_of_items[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Package Quantity": "item_package_quantity[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Color": "color[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Size": "size[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Theme": "theme[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Weave Type": "weave_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Care Instructions": "care_instructions[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Towel Form Type": "towel_form_type[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Weight Unit": "item_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Item Length Longer Edge": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Item Length Unit": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Width Shorter Edge": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Item Width Unit": "item_length_width[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Package Length": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.value",
        "Package Length Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.length.unit",
        "Item Package Width": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.value",
        "Package Width Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.width.unit",
        "Item Package Height": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.value",
        "Package Height Unit": "item_package_dimensions[marketplace_id=ATVPDKIKX0DER]#1.height.unit",
        "Package Weight": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Package Weight Unit": "item_package_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Item Display Weight": "item_display_weight[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Item Display Weight Unit": "item_display_weight[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Country of Origin": "country_of_origin[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries required?": "batteries_required[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Are batteries included?": "batteries_included[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Dangerous Goods Regulations": "supplier_declared_dg_hz_regulation[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Product Compliance Certificate": "required_product_compliance_certificate[marketplace_id=ATVPDKIKX0DER]#1.value",
    },
    "PROTECTIVE_GLOVE": {
        "Model Name": "model_name[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Style": "style[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Department Name": "department[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Age Range Description": "age_range_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Fabric Type": "fabric_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Fit Type": "fit_type[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Import Designation": "import_designation[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
        "Item Thickness Decimal Value": "item_thickness[marketplace_id=ATVPDKIKX0DER]#1.decimal_value",
        "Item Thickness": "item_thickness[marketplace_id=ATVPDKIKX0DER]#1.string_value",
        "Item Thickness Unit": "item_thickness[marketplace_id=ATVPDKIKX0DER]#1.unit",
        "Warranty Description": "warranty_description[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value",
    }
}

PRODUCT_TYPE_DISALLOWED_FIELDS = {
    "TOWEL": {
        "Compliance - Towel End Use": "compliance_towel_end_use[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Compliance Weave Type": "compliance_weave_type[marketplace_id=ATVPDKIKX0DER]#1.value",
        "Compliance - Outer Surface Material": "compliance_outer_surface_material[marketplace_id=ATVPDKIKX0DER]#1.value",
    }
}


def _required_fields_from_data_definitions(wb):
    if "Data Definitions" not in wb.sheetnames:
        return {}
    ws = wb["Data Definitions"]
    required = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        values = [str(value).strip() if value is not None else "" for value in row]
        if len(values) < 6:
            continue
        field_name = values[1]
        label = values[2] or field_name
        if field_name and values[5].lower() == "required":
            required[field_name] = label
    return required


def _is_parent_optional_required_field(field_name):
    lowered = field_name.lower()
    return any(token in lowered for token in [
        "color[",
        "size[",
        "title_differentiation",
        "condition_type",
        "model_number",
        "model_name",
        "manufacturer",
        "part_number",
        "special_feature",
        "mounting_type",
        "required_product_compliance_certificate",
        "list_price",
        "purchasable_offer",
        "skip_offer",
        "main_product_image",
        "item_package_dimensions",
        "item_package_weight",
        "item_depth_width_height",
        "item_length_width_height",
        "item_length_width",
        "item_dimensions",
        "item_weight",
        "is_heat_sensitive",
        "is_expiration_dated_product",
        "contains_liquid_contents",
        "unit_count",
        "number_of_items",
        "number_of_packs",
        "number_of_boxes",
        "item_package_quantity",
        "included_components",
        "specific_uses_for_product",
        "recommended_uses_for_product",
        "indoor_outdoor_usage",
        "pet_toy_type",
        "pet_type",
        "breed_recommendation",
        "directions",
        "theme",
    ])


def validate_template_file(path, output_path=None, write_report=False):
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = find_template_sheet(wb)
    if ws is None:
        raise ValueError(f"文件没有 {template_sheet_names_text()} 页：{path}")
    field_to_col = {
        str(ws.cell(5, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(5, col).value
    }

    findings = []
    data_rows = []
    sku_col = field_to_col.get(FIELD_NAMES["sku"])
    if not sku_col:
        findings.append(error("-", "SKU", "找不到 SKU 字段列。", "确认第 5 行字段名没有被改动。"))
    else:
        for row in range(7, ws.max_row + 1):
            if ws.cell(row, sku_col).value not in (None, ""):
                data_rows.append(row)

    condition_col = field_to_col.get(FIELD_NAMES["condition"])
    item_highlight_col = field_to_col.get(FIELD_NAMES["item_highlight"])
    skip_offer_col = field_to_col.get(FIELD_NAMES["skip_offer"])
    list_price_col = field_to_col.get(FIELD_NAMES["list_price"])
    haul_price_col = field_to_col.get(FIELD_NAMES["haul_price"])
    parentage_col = field_to_col.get(FIELD_NAMES["parentage_level"])
    parent_sku_col = field_to_col.get(FIELD_NAMES["parent_sku"])
    variation_theme_col = field_to_col.get(FIELD_NAMES["variation_theme"])
    title_col = field_to_col.get(FIELD_NAMES["title"])
    color_col = field_to_col.get(FIELD_NAMES["color"])
    size_col = field_to_col.get(FIELD_NAMES["size"])
    number_of_items_col = field_to_col.get(FIELD_NAMES["number_of_items"])
    item_package_quantity_col = field_to_col.get(FIELD_NAMES["item_package_quantity"])
    contains_battery_or_cell_col = field_to_col.get(FIELD_NAMES["contains_battery_or_cell"])
    dog_breed_size_col = field_to_col.get(FIELD_NAMES["dog_breed_size"])
    product_type_col = field_to_col.get("product_type#1.value")
    required_fields = _required_fields_from_data_definitions(wb)
    dimension_pairs = [
        ("Item Depth", field_to_col.get(FIELD_NAMES["item_depth"]), field_to_col.get(FIELD_NAMES["item_depth_unit"])),
        ("Item Height", field_to_col.get(FIELD_NAMES["item_height"]), field_to_col.get(FIELD_NAMES["item_height_unit"])),
        ("Item Width", field_to_col.get(FIELD_NAMES["item_width"]), field_to_col.get(FIELD_NAMES["item_width_unit"])),
    ]
    available_dimension_pairs = [pair for pair in dimension_pairs if pair[1] and pair[2]]
    row_infos = {}
    variation_theme_values_by_product_type = {}

    for row in data_rows:
        row_infos[row] = {
            "row": row,
            "sku": ws.cell(row, sku_col).value if sku_col else "",
            "parentage": ws.cell(row, parentage_col).value if parentage_col else "",
            "parent_sku": ws.cell(row, parent_sku_col).value if parent_sku_col else "",
            "variation_theme": ws.cell(row, variation_theme_col).value if variation_theme_col else "",
            "title": ws.cell(row, title_col).value if title_col else "",
            "color": ws.cell(row, color_col).value if color_col else "",
            "size": ws.cell(row, size_col).value if size_col else "",
            "number_of_items": ws.cell(row, number_of_items_col).value if number_of_items_col else "",
            "package_quantity": ws.cell(row, item_package_quantity_col).value if item_package_quantity_col else "",
        }

    for row in data_rows:
        sku = ws.cell(row, sku_col).value
        parentage = str(ws.cell(row, parentage_col).value or "").strip().lower() if parentage_col else ""
        is_parent = parentage == "parent"
        is_child = parentage == "child"
        if condition_col and not is_parent:
            condition = ws.cell(row, condition_col).value
            if condition != "New":
                findings.append(error(row, "Item Condition", f"{sku} 的 Item Condition 不是 New，当前为 `{condition}`。", "新品统一填写 New。"))
        elif not condition_col:
            findings.append(error(row, "Item Condition", "模板中找不到 Item Condition 字段。", "确认模板是否包含 condition_type 字段。"))

        if skip_offer_col:
            skip_offer = ws.cell(row, skip_offer_col).value
            if skip_offer not in (None, ""):
                findings.append(error(row, "Skip Offer", f"{sku} 的 Skip Offer 应留空，当前为 `{skip_offer}`。", "不要填写 skip_offer，让报价随价格字段正常生成。"))

        if contains_battery_or_cell_col:
            contains_battery_or_cell = ws.cell(row, contains_battery_or_cell_col).value
            if contains_battery_or_cell == "No":
                findings.append(error(row, "Contains Battery or Cell", f"{sku} 的 Contains Battery or Cell 不能填 No。", "该字段有效值是 Battery 或 Cell；非电池商品按当前规则留空。"))

        if dog_breed_size_col:
            dog_breed_size = ws.cell(row, dog_breed_size_col).value
            if dog_breed_size == "All Breed Sizes":
                findings.append(error(row, "Dog Breed Size", f"{sku} 的 Dog Breed Size 不能填 All Breed Sizes。", "ANIMAL_COLLAR 模板有效值为 Extra Small、Small、Medium、Large、Giant、All；通用值填 All。"))

        if list_price_col and haul_price_col and not is_parent:
            list_price = ws.cell(row, list_price_col).value
            haul_price = ws.cell(row, haul_price_col).value
            if list_price in (None, ""):
                findings.append(error(row, "List Price", f"{sku} 的 List Price 为空。", "上传前填写数字价格。"))
            if haul_price in (None, ""):
                findings.append(error(row, "Haul Price", f"{sku} 的 Haul Price 为空。", "Haul/BZR 价格应与 List Price 同步。"))
        if item_highlight_col:
            title_col = field_to_col.get(FIELD_NAMES["title"])
            title = str(ws.cell(row, title_col).value or "") if title_col else ""
            item_highlight = ws.cell(row, item_highlight_col).value
            if item_highlight not in (None, "") and len(title) > 75:
                findings.append(error(row, "Item Highlight", f"{sku} 填写了 Item Highlight，但 Item Name 长度为 {len(title)}，超过 75 字符。", "标题按 100-125 字符规则时清空 Item Highlight；只有标题不超过 75 字符时才填写该字段。"))

        for label, value_col, unit_col in available_dimension_pairs:
            value = ws.cell(row, value_col).value
            unit = ws.cell(row, unit_col).value
            if value not in (None, "") and unit in (None, ""):
                findings.append(error(row, f"{label} Unit", f"{sku} 的 {label} 有数值但单位为空。", "尺寸数值和单位必须成对填写。"))
            if value in (None, "") and unit not in (None, ""):
                findings.append(error(row, label, f"{sku} 的 {label} 单位有值但数值为空。", "尺寸数值和单位必须成对填写。"))

        if is_parent:
            parent_sku = ws.cell(row, parent_sku_col).value if parent_sku_col else None
            variation_theme = ws.cell(row, variation_theme_col).value if variation_theme_col else None
            if parent_sku not in (None, ""):
                findings.append(error(row, "Parent SKU", f"{sku} 是 Parent 行，不应填写 Parent SKU `{parent_sku}`。", "清空 Parent 行的 child_parent_sku_relationship.parent_sku。"))
            if variation_theme in (None, ""):
                findings.append(error(row, "Variation Theme", f"{sku} 是 Parent 行，但 Variation Theme 为空。", "Parent 行填写 Color、Size 或 ColorSize。"))
        if is_child:
            parent_sku = ws.cell(row, parent_sku_col).value if parent_sku_col else None
            variation_theme = ws.cell(row, variation_theme_col).value if variation_theme_col else None
            if parent_sku in (None, ""):
                findings.append(error(row, "Parent SKU", f"{sku} 是 Child 行，但 Parent SKU 为空。", "Child 行填写对应父 SKU。"))
            if variation_theme in (None, ""):
                findings.append(error(row, "Variation Theme", f"{sku} 是 Child 行，但 Variation Theme 为空。", "Child 行 Variation Theme 与 Parent 保持一致。"))

        for label, field_name in COPY_FIELD_NAMES.items():
            col = field_to_col.get(field_name)
            if not col:
                continue
            value = ws.cell(row, col).value
            if CJK_RE.search(str(value or "")):
                findings.append(error(row, label, f"{sku} 的 {label} 包含中文。", "改成自然的跨境英语表达，不要把中文原文写入 Amazon 模板。"))

        listing_row = {}
        for key in ["title", "description", "generic_keyword"]:
            field_name = FIELD_NAMES.get(key)
            col = field_to_col.get(field_name) if field_name else None
            if col:
                listing_row[key] = ws.cell(row, col).value
        for idx in range(1, 6):
            field_name = COPY_FIELD_NAMES[f"Bullet {idx}"]
            col = field_to_col.get(field_name)
            if col:
                listing_row[f"bullet_{idx}"] = ws.cell(row, col).value
        material_col = field_to_col.get("material[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value")
        if material_col:
            listing_row["material"] = ws.cell(row, material_col).value
        for field, message, fix in validate_listing_row(listing_row):
            findings.append(error(row, field, f"{sku} 的 {message}", fix))

        if product_type_col:
            product_type = ws.cell(row, product_type_col).value
            product_type_text = _text(product_type)
            if product_type_text not in variation_theme_values_by_product_type:
                variation_theme_values_by_product_type[product_type_text] = _variation_theme_values(wb, product_type_text)
            allowed_variation_themes = variation_theme_values_by_product_type[product_type_text]
            if variation_theme_col and allowed_variation_themes:
                variation_theme = ws.cell(row, variation_theme_col).value
                if variation_theme not in (None, "") and _text(variation_theme) not in allowed_variation_themes:
                    findings.append(error(row, "Variation Theme", f"{sku} 的 Variation Theme `{variation_theme}` 不在当前模板 Valid Values 中。", "按当前模板 Valid Values 精确填写，包括大小写、斜杠和空格，例如 COLOR。"))
            for field_name, label in required_fields.items():
                col = field_to_col.get(field_name)
                if not col:
                    continue
                if is_parent and _is_parent_optional_required_field(field_name):
                    continue
                value = ws.cell(row, col).value
                if value in (None, ""):
                    findings.append(error(row, label, f"{sku} 的 {label} 为空。", "该字段在模板 Data Definitions 中标记为 Required，上传前应补齐。"))
            for label, field_name in PRODUCT_TYPE_CONDITIONAL_FIELDS.get(str(product_type), {}).items():
                col = field_to_col.get(field_name)
                if not col:
                    continue
                if is_parent and _is_parent_optional_required_field(field_name):
                    continue
                value = ws.cell(row, col).value
                if value in (None, ""):
                    findings.append(error(row, label, f"{sku} 的 {label} 为空。", f"{product_type} 模板上传前应补齐该条件必填字段。"))
            for label, field_name in PRODUCT_TYPE_DISALLOWED_FIELDS.get(str(product_type), {}).items():
                col = field_to_col.get(field_name)
                if not col:
                    continue
                value = ws.cell(row, col).value
                if value not in (None, ""):
                    findings.append(error(row, label, f"{sku} 的 {label} 当前不应填写，值为 `{value}`。", "TOWEL 当前 ships_globally 条件下该合规字段不允许提交，清空该字段。"))

    _validate_variation_titles(findings, row_infos)

    if write_report:
        if output_path is None:
            output_path = OUTPUTS_DIR / f"{path.stem}_模板自检报告.md"
        else:
            output_path = Path(output_path)
        write_template_report(output_path, path, findings, data_rows)
    else:
        output_path = None
    return findings, output_path


def error(row, field, message, fix):
    return {
        "severity": "error",
        "row": row,
        "field": field,
        "message": message,
        "fix": fix,
    }


def write_template_report(path, checked_file, findings, data_rows):
    errors = sum(1 for item in findings if item["severity"] == "error")
    lines = [
        f"# {Path(checked_file).name} 模板自检报告",
        "",
        f"- SKU 行数：{len(data_rows)}",
        f"- Error：{errors}",
        "",
    ]
    if not findings:
        lines.extend([
            "未发现模板级问题。",
            "",
            "已确认：",
            "",
            "- 非 Parent 行 Item Condition = New，Parent 行可留空",
            "- Skip Offer 留空",
            "- List Price / Haul Price 已填写",
            "- 卖方最低价格默认留空，不限制人工填写值",
            "- 标题超过 75 字符时 Item Highlight 留空",
            "- 模板包含 Item Depth/Height/Width 字段时，其数值与单位成对填写",
            ""
        ])
    else:
        lines.extend(["## 问题清单", ""])
        for item in findings:
            lines.extend([
                f"### [ERROR] 行 {item['row']} - {item['field']}",
                "",
                f"- 问题：{item['message']}",
                f"- 建议：{item['fix']}",
                "",
            ])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
